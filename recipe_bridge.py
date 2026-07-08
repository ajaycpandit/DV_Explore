"""
recipe_bridge.py — parse and render DeltaV BATCH_RECIPE procedures.

A BATCH_RECIPE export contains:
  - recipe metadata (author, product, batch sizes, description, version)
  - FORMULA_PARAMETER entries (recipe-level inputs, grouped Recipe/Report/…)
  - UNIT_ALIAS entries (which units the procedure binds to)
  - a PFC_ALGORITHM: the procedure function chart — an SFC at the procedure
    level whose STEPs reference unit procedures / operations, connected by
    TRANSITIONs (with expressions) via STEP_TRANSITION_CONNECTION and
    TRANSITION_STEP_CONNECTION edges.

This module is additive and does not touch core/. It reuses db_parser.extract_block
for brace matching.
"""

import re
import html

import db_parser


def _blk(text, start_brace_search_from):
    return db_parser.extract_block(text, text.index('{', start_brace_search_from))


def edit_formula_values(text, recipe_name, formula_name, changes):
    """Minimal-diff write-back of formula parameter values. `changes` is
    {param_name: new_value}. Only the CV= (or STRING_CV=) token inside the matching
    ATTRIBUTE_INSTANCE of the named BATCH_RECIPE_FORMULA is replaced — everything else
    stays byte-identical, so the result is a provable minimal diff of the original
    export that DeltaV can re-import. Returns (new_text, applied_list, skipped_list)."""
    # locate the specific formula block
    applied, skipped = [], []
    fm = None
    for m in re.finditer(r'BATCH_RECIPE_FORMULA\s+NAME="([^"]+)"\s+RECIPE="([^"]+)"', text):
        if m.group(1) == formula_name and m.group(2) == recipe_name:
            fm = m
            break
    if fm is None:
        return text, [], list(changes.keys())
    start = text.index('{', fm.start())
    blk = db_parser.extract_block(text, start)
    blk_start = start
    blk_end = start + len(blk)
    new_blk = blk
    # for each requested change, rewrite the CV inside that param's ATTRIBUTE_INSTANCE
    for pname, newval in changes.items():
        # match the attribute instance for this parameter within the block
        am = re.search(
            r'(ATTRIBUTE_INSTANCE\s+NAME="' + re.escape(pname) + r'"\s*\{\s*VALUE\s*\{\s*)'
            r'(CV=)([^\s}]+)', new_blk)
        if am:
            new_blk = new_blk[:am.start(3)] + str(newval) + new_blk[am.end(3):]
            applied.append(pname)
            continue
        sm = re.search(
            r'(ATTRIBUTE_INSTANCE\s+NAME="' + re.escape(pname) + r'"\s*\{\s*VALUE\s*\{\s*)'
            r'(STRING_CV=")([^"]*)(")', new_blk)
        if sm:
            new_blk = new_blk[:sm.start(3)] + str(newval) + new_blk[sm.end(3):]
            applied.append(pname)
            continue
        skipped.append(pname)
    new_text = text[:blk_start] + new_blk + text[blk_end:]
    return new_text, applied, skipped


def parse_formulas(text):
    """Parse BATCH_RECIPE_FORMULA blocks — the named value sets (e.g. 'MEDIA AND 20')
    that populate a recipe's parameters with actual literals. Returns
    {recipe_name: [{'name':..., 'description':..., 'values': {param: value}}]}."""
    out = {}
    for m in re.finditer(r'BATCH_RECIPE_FORMULA\s+NAME="([^"]+)"\s+RECIPE="([^"]+)"', text):
        try:
            blk = db_parser.extract_block(text, text.index('{', m.start()))
        except Exception:
            continue
        fname, rname = m.group(1), m.group(2)
        desc = re.search(r'DESCRIPTION="([^"]*)"', blk)
        values = {}
        for am in re.finditer(r'ATTRIBUTE_INSTANCE\s+NAME="([^"]+)"\s*\{\s*VALUE\s*\{\s*([^}]*)\}', blk):
            pname, vbody = am.group(1), am.group(2)
            cv = re.search(r'\bCV=([^\s}]+)', vbody)
            scv = re.search(r'STRING_CV="([^"]*)"', vbody)
            if cv:
                values[pname] = cv.group(1)
            elif scv:
                values[pname] = scv.group(1)
        out.setdefault(rname, []).append({
            'name': fname,
            'description': desc.group(1) if desc else '',
            'released': 'RELEASED_TO_PRODUCTION=T' in blk,
            'values': values,
        })
    return out


def parse_recipes(text):
    """Parse ALL BATCH_RECIPE blocks in the export into a list of structured dicts.
    Excludes BATCH_RECIPE_FORMULA (a different block type — the named value sets
    handled by parse_formulas) via the (?!_) guard, since a plain regex boundary
    doesn't stop at a literal underscore."""
    out = []
    for m in re.finditer(r'BATCH_RECIPE(?!_)\s+NAME="([^"]+)"(?:\s+TYPE=(\w+))?', text):
        try:
            blk = db_parser.extract_block(text, text.index('{', m.start()))
        except Exception:
            continue
        name = m.group(1)
        rtype = m.group(2) or ''
        out.append(_parse_one_recipe(blk, name, rtype))
    return out


def parse_recipe(text):
    """Back-compat: parse the FIRST BATCH_RECIPE, or None if absent."""
    recs = parse_recipes(text)
    return recs[0] if recs else None


def categorize(recipes):
    """Group parsed recipes into the three DeltaV categories by BATCH_RECIPE TYPE:
    Procedure / Unit Procedure / Operation. Returns an ordered dict-like list of
    (category, [recipes])."""
    cats = {'Procedure': [], 'Unit Procedure': [], 'Operation': []}
    label = {'PROCEDURE': 'Procedure', 'UNIT_PROCEDURE': 'Unit Procedure',
             'OPERATION': 'Operation'}
    for r in recipes:
        t = (r['meta'].get('type') or 'PROCEDURE').upper()
        cats.setdefault(label.get(t, 'Procedure'), []).append(r)
    return cats


def all_step_children(recipes):
    """Map each recipe name -> ordered list of {step, name, layer, loaded} for every
    step instance it references (e.g. CENT_HC_INIT_UP:1 and :2 listed separately,
    since each instance carries its own deferral set) — even when the referenced
    object hasn't been imported yet. 'name' is the base definition; 'step' the
    instance name shown in the tree."""
    known = set(r['meta']['name'] for r in recipes)
    out = {}
    for r in recipes:
        proc = r.get('procedure') or {}
        kids = []
        for sn, s in (proc.get('steps') or {}).items():
            if sn in ('START', 'END'):
                continue
            base = re.sub(r':\d+$', '', s.get('definition', ''))
            if not base or base == r['meta']['name']:
                continue
            kids.append({'step': sn, 'name': base,
                        'layer': _layer_label(s.get('definition', '')),
                        'loaded': base in known,
                        'has_params': bool(s.get('params'))})
        out[r['meta']['name']] = kids
    return out


def child_links(recipes):
    """Map each recipe name -> list of (step_name, child_recipe_name) for steps whose
    definition matches another recipe object (the drill-down / tree-nesting relation)."""
    names = set(r['meta']['name'] for r in recipes)
    links = {}
    for r in recipes:
        proc = r.get('procedure') or {}
        kids = []
        for sn, s in (proc.get('steps') or {}).items():
            base = re.sub(r':\d+$', '', s.get('definition', ''))
            if base in names and base != r['meta']['name']:
                kids.append((sn, base))
        links[r['meta']['name']] = kids
    return links


def resolve_deferrals(recipe, parent_recipe=None):
    """For a recipe's steps, resolve each deferred parameter's value. A deferred
    param on a step (targeting the child object) has DEFERRED_TO pointing at a
    parameter *at this recipe's level* (a formula parameter) that supplies the value.
    Returns {step_name: [{name, deferred_to, resolved_value, source}]}. If the
    referenced parameter itself is deferred at the parent level, we surface the
    parent parameter name (the chain continues one level up)."""
    my_params = {p['name']: p for p in recipe.get('params', [])}
    out = {}
    proc = recipe.get('procedure') or {}
    for sn, s in (proc.get('steps') or {}).items():
        rows = []
        for d in s.get('deferred', []):
            dto = d.get('deferred_to') or d.get('name')
            src = my_params.get(dto)
            if src and src.get('value'):
                rows.append({'name': d['name'], 'deferred_to': dto,
                             'resolved_value': src['value'], 'source': 'this level'})
            elif src:
                # exists here but no literal value -> passed down from above
                rows.append({'name': d['name'], 'deferred_to': dto,
                             'resolved_value': '(deferred to ' + dto + ')',
                             'source': 'parent level'})
            else:
                rows.append({'name': d['name'], 'deferred_to': dto,
                             'resolved_value': '(unresolved)', 'source': ''})
        if rows:
            out[sn] = rows
    return out


def _parse_one_recipe(blk, name, rtype=''):
    """Parse a single BATCH_RECIPE block body into a structured dict."""

    def _scalar(key):
        mm = re.search(key + r'="([^"]*)"', blk)
        if mm:
            return mm.group(1)
        mm = re.search(key + r'=([^\s{][^\r\n]*)', blk)
        return mm.group(1).strip() if mm else ''

    # derive name from DESCRIPTION or a NAME= if not on the header
    if not name:
        nm = re.search(r'\bNAME="([^"]+)"', blk[:200])
        name = nm.group(1) if nm else (_scalar('DESCRIPTION')[:30] or 'RECIPE')

    # BATCH_RECIPE TYPE (from the header line) tells us the layer:
    # PROCEDURE / UNIT_PROCEDURE / OPERATION.
    if not rtype:
        thdr = re.search(r'^\s*TYPE=(\w+)', blk[:400], re.MULTILINE)
        rtype = thdr.group(1) if thdr else 'PROCEDURE'
    meta = {
        'name': name,
        'type': rtype,
        'description': _scalar('DESCRIPTION'),
        'author': _scalar('AUTHOR'),
        'abstract': _scalar('ABSTRACT'),
        'product_code': _scalar('PRODUCT_CODE'),
        'product_name': _scalar('PRODUCT_NAME'),
        'version': _scalar('VERSION'),
        'batch_units': _scalar('BATCH_UNITS'),
        'default_batch_size': _scalar('DEFAULT_BATCH_SIZE'),
        'min_batch_size': _scalar('MINIMUM_BATCH_SIZE'),
        'max_batch_size': _scalar('MAXIMUM_BATCH_SIZE'),
        'approval': _scalar('RECIPE_APPROVAL_INFO'),
    }

    params = []
    for pm in re.finditer(r'FORMULA_PARAMETER\s+NAME="([^"]+)"\s+TYPE=(\w+)\s*\{', blk):
        pbody = db_parser.extract_block(blk, pm.end() - 1)
        grp = re.search(r'GROUP="([^"]*)"', pbody)
        conn = re.search(r'CONNECTION=(\w+)', pbody)
        desc = re.search(r'DESCRIPTION="([^"]*)"', pbody)
        val = re.search(r'\bVALUE="([^"]*)"', pbody) or re.search(r'\bVALUE=([^\s{][^\r\n]*)', pbody)
        strval = re.search(r'STRING_VALUE="([^"]*)"', pbody)
        setref = re.search(r'\bSET="([^"]*)"', pbody)
        locked = 'IS_PARAMETER_LOCKED=T' in pbody
        params.append({
            'name': pm.group(1), 'type': pm.group(2),
            'group': grp.group(1) if grp else '',
            'connection': conn.group(1) if conn else '',
            'description': desc.group(1) if desc else '',
            'value': (val.group(1).strip() if val else (strval.group(1) if strval else '')),
            'set': setref.group(1) if setref else '',
            'param_type': _fhx_param_type(pm.group(2)),
            'filtering': _fhx_filtering(conn.group(1) if conn else ''),
            'locked': locked,
        })

    aliases = []
    for am in re.finditer(r'UNIT_ALIAS\s+NAME="([^"]+)"\s*\{', blk):
        abody = db_parser.extract_block(blk, am.end() - 1)
        desc = re.search(r'DESCRIPTION="([^"]*)"', abody)
        aliases.append({'name': am.group(1),
                        'desc': desc.group(1) if desc else ''})

    proc = _parse_pfc(blk)
    return {'meta': meta, 'params': params, 'aliases': aliases, 'procedure': proc}


def _parse_pfc(blk):
    """Parse the PFC_ALGORITHM into steps, transitions, and edges (with coords)."""
    pm = re.search(r'PFC_ALGORITHM\s*\{', blk)
    if not pm:
        return None
    pfc = db_parser.extract_block(blk, pm.end() - 1)

    steps = {}
    for sm in re.finditer(r'(INITIAL_)?STEP\s+NAME="([^"]+)"\s+DEFINITION="([^"]+)"\s*\{', pfc):
        sname, sdef = sm.group(2), sm.group(3)
        sbody = db_parser.extract_block(pfc, sm.end() - 1)
        desc = re.search(r'DESCRIPTION="([^"]*)"', sbody)
        rect = re.search(r'RECTANGLE=\s*\{\s*X=(-?\d+)\s+Y=(-?\d+)\s+H=(\d+)\s+W=(\d+)', sbody)
        ualias = re.search(r'UNIT_ALIAS="([^"]*)"', sbody)
        sparams = []
        deferred = []
        for spm in re.finditer(r'STEP_PARAMETER\s+NAME="([^"]+)"\s*\{', sbody):
            spb = db_parser.extract_block(sbody, spm.end() - 1)
            origin = re.search(r'ORIGIN=(\w+)', spb)
            grp = re.search(r'GROUP="([^"]*)"', spb)
            dto = re.search(r'DEFERRED_TO="([^"]*)"', spb)
            val = re.search(r'VALUE="([^"]*)"', spb) or re.search(r'VALUE=([^\s{][^\r\n]*)', spb)
            entry = {'name': spm.group(1),
                     'origin': origin.group(1) if origin else '',
                     'group': grp.group(1) if grp else '',
                     'deferred_to': dto.group(1) if dto else '',
                     'value': (val.group(1).strip() if val else '')}
            sparams.append(entry)
            if entry['origin'] == 'DEFERRED':
                deferred.append(entry)
        steps[sname] = {
            'name': sname, 'definition': sdef,
            'desc': desc.group(1) if desc else '',
            'initial': bool(sm.group(1)),
            'params': sparams,
            'deferred': deferred,
            'unit_alias': ualias.group(1) if ualias else '',
            'x': int(rect.group(1)) if rect else None,
            'y': int(rect.group(2)) if rect else None,
            'h': int(rect.group(3)) if rect else 40,
            'w': int(rect.group(4)) if rect else 140,
        }

    transitions = {}
    for tm in re.finditer(r'TRANSITION\s+NAME="([^"]+)"\s*\{', pfc):
        tbody = db_parser.extract_block(pfc, tm.end() - 1)
        # FHX escapes a literal " inside a quoted string by DOUBLING it (""), not with
        # a backslash. Use a doubled-quote-aware capture so expressions that compare
        # against a string literal (e.g. != "") aren't truncated at the first embedded
        # quote — the same fix applied to phase SFC transitions via sfc_expr_fix.
        expr = re.search(r'EXPRESSION="((?:[^"]|"")*)"', tbody, re.DOTALL)
        desc = re.search(r'DESCRIPTION="((?:[^"]|"")*)"', tbody, re.DOTALL)
        _expr = (expr.group(1) if expr else '').replace('""', '"')
        transitions[tm.group(1)] = {
            'name': tm.group(1),
            'expr': _expr.replace('\r\n', ' ').replace('\n', ' ').strip(),
            'desc': (desc.group(1) if desc else '').replace('""', '"'),
        }

    s2t, t2s = [], []
    for cm in re.finditer(r'STEP_TRANSITION_CONNECTION\s+STEP="([^"]+)"\s+TRANSITION="([^"]+)"', pfc):
        s2t.append((cm.group(1), cm.group(2)))
    for cm in re.finditer(r'TRANSITION_STEP_CONNECTION\s+TRANSITION="([^"]+)"\s+STEP="([^"]+)"', pfc):
        t2s.append((cm.group(1), cm.group(2)))

    return {'steps': steps, 'transitions': transitions, 's2t': s2t, 't2s': t2s}


def _render_pfc_svg(proc, known_recipes=None):
    known_recipes = known_recipes or set()
    """Render the procedure as an SVG SFC diagram using the step X/Y coordinates.
    Transitions are placed on the edges between their source and target steps."""
    steps = proc['steps']
    trans = proc['transitions']
    s2t, t2s = proc['s2t'], proc['t2s']

    # step positions; some steps (START pseudo) have none — synthesize a top slot
    coords = {n: (s['x'], s['y'], s['w'], s['h']) for n, s in steps.items()
              if s['x'] is not None}
    if not coords:
        return ''  # no coordinates, can't draw

    # index edges
    step_outs = {}
    for s, t in s2t:
        step_outs.setdefault(s, []).append(t)
    trans_outs = {}
    for t, s in t2s:
        trans_outs.setdefault(t, []).append(s)
    trans_in = {}
    for s, t in s2t:
        trans_in.setdefault(t, []).append(s)

    # place each transition at the midpoint between its source step(s) and target(s)
    tpos = {}
    for tn in trans:
        srcs = [coords[s] for s in trans_in.get(tn, []) if s in coords]
        tgts = [coords[s] for s in trans_outs.get(tn, []) if s in coords]
        pts = srcs + tgts
        if pts:
            cx = sum(p[0] + p[2] / 2 for p in pts) / len(pts)
            cy = sum(p[1] + p[3] / 2 for p in pts) / len(pts)
            tpos[tn] = (cx, cy)

    # compute viewbox
    xs = [c[0] for c in coords.values()] + [c[0] + c[2] for c in coords.values()]
    ys = [c[1] for c in coords.values()] + [c[1] + c[3] for c in coords.values()]
    pad = 40
    minx, maxx = min(xs) - pad, max(xs) + pad
    miny, maxy = min(ys) - pad, max(ys) + pad
    W, H = maxx - minx, maxy - miny

    def sx(x):
        return x - minx

    def sy(y):
        return y - miny

    svg = [f'<svg class="pfc-svg" viewBox="0 0 {W} {H}" '
           f'xmlns="http://www.w3.org/2000/svg" style="min-width:{min(W, 900)}px">']
    svg.append('<defs><marker id="pfcarr" markerWidth="8" markerHeight="8" refX="6" refY="4" '
               'orient="auto"><path d="M0,0 L8,4 L0,8 z" fill="#94a3b8"/></marker></defs>')

    # edges: step -> transition -> step
    def center(name):
        x, y, w, h = coords[name]
        return sx(x) + w / 2, sy(y) + h / 2

    def ortho(x1, y1, x2, y2, arrow=False):
        # L-shaped connector: vertical from source, then horizontal to target — matches
        # the phase SFC's straight-line style (#5) rather than a diagonal.
        mid = f'{x1:.0f},{y1:.0f} {x1:.0f},{y2:.0f} {x2:.0f},{y2:.0f}'
        mk = ' marker-end="url(#pfcarr)"' if arrow else ''
        return (f'<polyline points="{mid}" fill="none" stroke="#94a3b8" '
                f'stroke-width="1.4"{mk}/>')

    for s, t in s2t:
        if s not in coords or t not in tpos:
            continue
        x1, y1 = center(s)
        x2, y2 = tpos[t][0] - minx, tpos[t][1] - miny
        svg.append(ortho(x1, y1, x2, y2))
    for t, s in t2s:
        if s not in coords or t not in tpos:
            continue
        x1, y1 = tpos[t][0] - minx, tpos[t][1] - miny
        x2, y2 = center(s)
        svg.append(ortho(x1, y1, x2, y2, arrow=True))

    # step boxes
    for n, (x, y, w, h) in coords.items():
        s = steps[n]
        fill = '#0f172a' if n == 'START' else '#ffffff'
        txtcol = '#fff' if n == 'START' else '#16202c'
        initial = s.get('initial')
        stroke = '#16a34a' if initial else '#334155'
        sw = 2.5 if initial else 1.5
        label = html.escape(n)
        defn = html.escape(s.get('definition', ''))
        base_def = re.sub(r':\d+$', '', s.get('definition', ''))
        drill = (' data-drill="' + html.escape(base_def, quote=True) + '"') if base_def in known_recipes else ''
        cls = 'pfc-step pfc-drillable' if drill else 'pfc-step'
        svg.append(f'<g class="{cls}" data-step="{html.escape(n, quote=True)}"{drill}>')
        svg.append(f'<rect x="{sx(x)}" y="{sy(y)}" width="{w}" height="{h}" rx="5" '
                   f'fill="{fill}" stroke="{stroke}" stroke-width="{sw}"/>')
        svg.append(f'<text x="{sx(x) + w/2:.0f}" y="{sy(y) + 16:.0f}" text-anchor="middle" '
                   f'font-size="11" font-weight="600" fill="{txtcol}">{label}</text>')
        if defn and n != 'START':
            svg.append(f'<text x="{sx(x) + w/2:.0f}" y="{sy(y) + 30:.0f}" text-anchor="middle" '
                       f'font-size="9" fill="#64748b">{defn[:22]}</text>')
        if drill:
            # marker showing this step has a child object to drill into
            svg.append(f'<text x="{sx(x) + w - 9:.0f}" y="{sy(y) + 13:.0f}" text-anchor="middle" '
                       f'font-size="11" fill="#7c3aed">\u25c9</text>')
        svg.append('</g>')

    # transition bars + labels
    for tn, (cx, cy) in tpos.items():
        px, py = cx - minx, cy - miny
        td = trans.get(tn, {})
        has_expr = bool(td.get('expr'))
        color = '#7c3aed' if has_expr else '#94a3b8'
        svg.append(f'<g class="pfc-trans" data-trans="{html.escape(tn, quote=True)}" '
                   f'style="cursor:pointer">')
        svg.append(f'<rect x="{px-30:.0f}" y="{py-7:.0f}" width="60" height="14" rx="3" '
                   f'fill="#fff" stroke="{color}" stroke-width="1.3"/>')
        svg.append(f'<text x="{px:.0f}" y="{py+3:.0f}" text-anchor="middle" font-size="9" '
                   f'font-weight="600" fill="{color}">{html.escape(tn[:12])}</text>')
        svg.append('</g>')

    svg.append('</svg>')
    return '\n'.join(svg)


# ───────────────────────── rendering ─────────────────────────

def build_recipe_html(recipe, known_recipes=None, deferral_map=None, formulas=None):
    known_recipes = known_recipes or set()
    deferral_map = deferral_map or {}
    # formulas: list of {name, description, values:{param:val}} — the named value sets
    # that populate the Parameter Value column; the first released one is the default.
    formulas = formulas or []
    """Render a parsed recipe as an HTML view: metadata + parameters + a
    procedure flow (steps and transitions with expressions)."""
    if not recipe:
        return '<span class="empty">No recipe in this export.</span>'
    meta = recipe['meta']
    h = []

    # procedure diagram (SVG) + flow (the main content)
    proc = recipe.get('procedure')
    if proc and proc['steps']:
        diagram = _render_pfc_svg(proc, known_recipes)
        h.append('<div class="card" style="max-width:none"><h3>Procedure diagram ('
                 + str(len(proc['steps'])) + ' steps, '
                 + str(len(proc['transitions'])) + ' transitions)</h3>')
        if diagram:
            # embed transition expressions as data attributes for click-to-view
            # (handled by the explorer's delegated handler — inline <script> injected
            # via innerHTML would never execute).
            import json as _json
            texpr = {tn: td.get('expr', '') for tn, td in proc['transitions'].items()}
            # #4: per-step detail (definition, params with deferred-to, child availability)
            # so clicking a step in the diagram shows its info in the side panel — the
            # same interaction the SFC phase view offers.
            sinfo = {}
            # step -> outgoing/incoming transitions (for sequence context in the panel)
            s2t_map = {}
            for a, b in (proc.get('s2t') or []):
                s2t_map.setdefault(a, []).append(b)
            t2s_map = {}
            for a, b in (proc.get('t2s') or []):
                t2s_map.setdefault(b, []).append(a)  # step b is fed by transition a
            for sn, s in proc['steps'].items():
                if sn in ('START', 'END'):
                    continue
                base_def = re.sub(r':\d+$', '', s.get('definition', ''))
                params = []
                n_def = 0
                for p in (s.get('params') or []):
                    isd = p.get('origin') == 'DEFERRED'
                    if isd:
                        n_def += 1
                    params.append({
                        'name': p.get('name', ''),
                        'deferred': p.get('deferred_to', '') if isd else '',
                        'group': p.get('group', ''),
                    })
                sinfo[sn] = {
                    'def': s.get('definition', ''),
                    'layer': _layer_label(s.get('definition', '')),
                    'params': params,
                    'n_def': n_def,
                    'outs': s2t_map.get(sn, []),
                    'ins': t2s_map.get(sn, []),
                    'initial': bool(s.get('initial')),
                    'child': base_def if (known_recipes and base_def in known_recipes) else '',
                }
            pid = 'pfcPanel_' + _safe_id(recipe['meta']['name'])
            h.append('<div class="pfc-toolbar">'
                     '<button class="pfc-zbtn" data-pfc-zoom="out" title="Zoom out">\u2212</button>'
                     '<button class="pfc-zbtn" data-pfc-zoom="reset" title="Reset">\u25a1</button>'
                     '<button class="pfc-zbtn" data-pfc-zoom="in" title="Zoom in">+</button>'
                     '<span class="pfc-hint2">click a step or transition for detail \u00b7 drag to pan \u00b7 scroll to zoom</span></div>')
            # #5: two-column layout — detail panel on the LEFT, diagram on the RIGHT,
            # so a selected step/transition's parameters sit beside the diagram instead
            # of pushing everything down.
            h.append('<div class="pfc-layout">')
            h.append('<div class="pfc-wrap" data-pfc-expr="'
                     + html.escape(_json.dumps(texpr), quote=True)
                     + '" data-pfc-steps="' + html.escape(_json.dumps(sinfo), quote=True)
                     + '" data-pfc-panel="' + pid + '"><div class="pfc-zoomlayer">' + diagram + '</div></div>')
            h.append('<div class="pfc-divider" title="Drag to resize"></div>')
            h.append('<div class="pfc-panel" id="' + pid
                     + '"><span class="pfc-hint">Click a step or transition in the diagram to see its detail. '
                     'Steps marked \u25c9 have a child object you can drill into (click the link, or right-click the step).</span></div>')
            h.append('</div>')
        else:
            h.append('<span class="empty">Diagram coordinates unavailable; see the flow below.</span>')
        h.append('</div>')

        # textual flow (steps + transitions interleaved) — starts collapsed (#4). Uses
        # the standard .card collapse mechanism (delegated h3 handler), not a custom
        # onclick, so it expands/collapses reliably in both Explorer and workspace.
        h.append('<div class="card collapsed" style="max-width:none"><h3>Procedure flow '
                 '<span style="font-weight:400;color:var(--ink-3);font-size:12px">'
                 '\u2014 full step &amp; transition listing</span></h3>')
        h.append(_render_flow(proc, known_recipes, {p['name']: p for p in recipe.get('params', [])}))
        h.append('</div>')

    # formula parameters — full grid matching the DeltaV Recipe Studio columns:
    # Name, Type, Description, Parameter Value, Filtering, Parameter Type, Group, Locked.
    params = recipe.get('params', [])
    if params:
        # build a per-formula value lookup embedded as data attributes so the selector
        # can switch the Parameter Value column without a round-trip.
        import json as _json
        fvals = {f['name']: f.get('values', {}) for f in formulas}
        default_formula = ''
        for f in formulas:
            if f.get('released'):
                default_formula = f['name']
                break
        if not default_formula and formulas:
            default_formula = formulas[0]['name']

        h.append('<div class="card" style="max-width:none"><h3>Parameters (' + str(len(params)) + ')</h3>')
        if formulas:
            opts = ''.join('<option value="' + html.escape(f['name'], quote=True) + '"'
                           + (' selected' if f['name'] == default_formula else '') + '>'
                           + html.escape(f['name']) + (' \u2713' if f.get('released') else '') + '</option>'
                           for f in formulas)
            h.append('<div class="rp-formula-bar">Formula: <select id="rpFormula" '
                     'data-fvals="' + html.escape(_json.dumps(fvals), quote=True) + '" '
                     'data-recipe="' + html.escape(recipe.get('name', ''), quote=True) + '" '
                     'onchange="rpApplyFormula(this)">' + opts + '</select>'
                     '<span class="rp-formula-hint">values shown come from the selected formula</span>'
                     '<span style="flex:1"></span>'
                     '<button class="rp-editbtn" onclick="rpToggleEdit(this)" '
                     'title="Edit this formula\'s values and export a minimal-diff FHX for re-import">'
                     '\u270e Edit values</button></div>')
        h.append('<input class="alias-filter" id="rpFilter" placeholder="Filter parameters\u2026" oninput="rpFilterGrid(this)">')
        h.append('<div class="rp-gridwrap"><table class="fbd-table rp-grid"><thead><tr>'
                 '<th>Name</th><th>Type</th><th>Description</th><th>Parameter Value</th>'
                 '<th>Filtering</th><th>Parameter Type</th><th>Group</th><th>Locked</th>'
                 '</tr></thead><tbody>')
        default_vals = fvals.get(default_formula, {})
        for p in params:
            # Parameter Value priority: selected formula's literal -> export literal ->
            # derived up-reference (parent param) -> em-dash.
            fv = default_vals.get(p['name'], '')
            pval = fv or p.get('value', '')
            if pval:
                pval_html = '<code class="rp-val">' + html.escape(pval) + '</code>'
            elif p['name'] in deferral_map:
                pval_html = ('<code class="rp-upref">' + html.escape(deferral_map[p['name']])
                             + '</code> <span class="rp-uparrow">\u2191 parent</span>')
            else:
                pval_html = '<span class="rp-none">\u2014</span>'
            h.append('<tr class="rp-row" data-pname="' + html.escape(p['name'], quote=True) + '"'
                     + (' data-upref="' + html.escape(deferral_map[p['name']], quote=True) + '"' if p['name'] in deferral_map else '')
                     + '>'
                     '<td><code>' + html.escape(p['name']) + '</code></td>'
                     '<td>' + _fhx_type_label(p.get('connection', '')) + '</td>'
                     '<td class="rp-desc">' + html.escape(p.get('description', '')) + '</td>'
                     '<td class="rp-valcell">' + pval_html + '</td>'
                     '<td>' + html.escape(p.get('filtering', '')) + '</td>'
                     '<td>' + html.escape(p.get('param_type', '')) + '</td>'
                     '<td>' + html.escape(p.get('group', '')) + '</td>'
                     '<td>' + ('\u2713' if p['locked'] else '') + '</td>'
                     '</tr>')
        h.append('</tbody></table></div></div>')

    # unit aliases
    aliases = recipe.get('aliases', [])
    if aliases:
        h.append('<div class="card"><h3>Unit aliases (' + str(len(aliases)) + ')</h3>'
                 '<table class="fbd-table"><thead><tr><th>Alias</th><th>Description</th></tr></thead><tbody>')
        for a in aliases:
            h.append('<tr><td><code>' + html.escape(a['name']) + '</code></td><td>'
                     + html.escape(a['desc']) + '</td></tr>')
        h.append('</tbody></table></div>')

    # deferrals audit — every step's parameters flattened, matching a manual
    # Recipe Studio deferral check, with a one-click Excel export.
    defer_rows = build_deferrals_rows(recipe)
    if defer_rows:
        h.append(build_deferrals_html(recipe, defer_rows))

    return '\n'.join(h)


def _render_flow(proc, known_recipes=None, my_params=None):
    """Render the procedure as steps interleaved with their outgoing transitions.
    known_recipes: set of recipe names present in the export, so a step whose
    definition references one becomes a drill-down link (#3).
    my_params: {name: param} for this recipe level, so a step's deferred parameter
    can show the value (or the parent parameter it defers to)."""
    known_recipes = known_recipes or set()
    my_params = my_params or {}
    steps, trans = proc['steps'], proc['transitions']
    s2t, t2s = proc['s2t'], proc['t2s']
    # index: step -> [transitions]; transition -> [steps]
    step_outs = {}
    for s, t in s2t:
        step_outs.setdefault(s, []).append(t)
    trans_outs = {}
    for t, s in t2s:
        trans_outs.setdefault(t, []).append(s)

    # order: start from initial step(s), walk breadth-first
    order = []
    seen = set()
    initials = [n for n, s in steps.items() if s.get('initial')]
    # 'START' is a pseudo-step referenced in edges but not a real step
    frontier = initials or list(steps.keys())[:1]
    # include START pseudo if present in edges
    if any(s == 'START' for s, _ in s2t):
        frontier = ['START'] + frontier
    while frontier:
        nxt = []
        for s in frontier:
            if s in seen:
                continue
            seen.add(s)
            order.append(s)
            for t in step_outs.get(s, []):
                for ns in trans_outs.get(t, []):
                    if ns not in seen:
                        nxt.append(ns)
        frontier = nxt
    # append any steps not reached
    for n in steps:
        if n not in seen:
            order.append(n)

    rows = ['<div class="rflow">']
    for sname in order:
        if sname == 'START':
            rows.append('<div class="rf-step rf-start">START</div>')
        else:
            s = steps.get(sname)
            if not s:
                continue
            label = html.escape(sname)
            defn = html.escape(s['definition'])
            npar = len(s.get('params', []))
            ndef = len(s.get('deferred', []))
            ualias = s.get('unit_alias', '')
            init = ' <span class="rf-init">initial</span>' if s.get('initial') else ''
            alias_tag = (' <span class="rf-alias">unit: ' + html.escape(ualias) + '</span>') if ualias else ''
            # if the definition matches another recipe object, make it a drill-down link
            base_def = re.sub(r':\d+$', '', s['definition'])
            if base_def in known_recipes:
                defn_html = ('<span class="rf-drill" data-recipe="' + html.escape(base_def, quote=True)
                             + '"><code>' + defn + '</code> \u2197</span>')
            else:
                defn_html = '<code>' + defn + '</code>'
            rows.append('<div class="rf-step"><div class="rf-name">' + label + init + alias_tag
                        + '</div><div class="rf-def">' + _layer_label(s['definition'])
                        + ': ' + defn_html
                        + (' \u00b7 ' + str(npar) + ' parameters' if npar else '')
                        + (' \u00b7 <span class="rf-defcount">' + str(ndef) + ' deferred</span>' if ndef else '')
                        + '</div>')
            # deferred parameters (the values passed down to the next layer)
            if s.get('deferred'):
                rows.append('<div class="rf-defers"><div class="rf-defhdr">Deferred parameters (value comes from the level above)</div>')
                for d in s['deferred']:
                    dto = d.get('deferred_to') or d['name']
                    src = my_params.get(dto)
                    if src and src.get('value'):
                        resolved = '<span class="rf-defval">= ' + html.escape(src['value']) + '</span>'
                    elif src:
                        resolved = '<span class="rf-defup">\u2191 ' + html.escape(dto) + '</span>'
                    else:
                        resolved = '<span class="rf-defun">unresolved</span>'
                    rows.append('<div class="rf-defrow"><code>' + html.escape(d['name']) + '</code>'
                                + ' \u2192 <code>' + html.escape(dto) + '</code> ' + resolved
                                + (' <span class="rf-defgrp">' + html.escape(d['group']) + '</span>' if d.get('group') else '')
                                + '</div>')
                rows.append('</div>')
            rows.append('</div>')  # close rf-step
        # outgoing transitions from this step
        for t in step_outs.get(sname, []):
            td = trans.get(t, {})
            expr = td.get('expr', '')
            targets = trans_outs.get(t, [])
            tgt = ', '.join(html.escape(x) for x in targets)
            rows.append('<div class="rf-trans"><span class="rf-tname">\u2500 ' + html.escape(t)
                        + '</span> <span class="rf-tgt">\u2192 ' + (tgt or '?') + '</span>'
                        + ('<div class="rf-expr">' + html.escape(expr) + '</div>' if expr else '')
                        + '</div>')
    rows.append('</div>')
    return '\n'.join(rows)


def build_step_param_html(recipe, step_name):
    """Build the parameter view for a child object instance (e.g. CENT_HC_INIT_UP:1)
    from the PARENT recipe's step data — usable before the child's own FHX is
    imported. The parent's STEP_PARAMETER entries are exactly the child's parameter
    list; a DEFERRED one shows the parent parameter supplying its value (the
    'Parameter Value' column in Recipe Studio), e.g. CENT1_SELECTED -> G007_SELECTED."""
    proc = recipe.get('procedure') or {}
    s = (proc.get('steps') or {}).get(step_name)
    if not s:
        return ''
    h = []
    ualias = s.get('unit_alias', '')
    h.append('<div class="card" style="max-width:none"><h3>Parameters ('
             + str(len(s.get('params', []))) + ')'
             + (' <span class="rf-alias">unit: ' + html.escape(ualias) + '</span>' if ualias else '')
             + '</h3>')
    h.append('<div class="rstep-note">Derived from step <code>' + html.escape(step_name)
             + '</code> of <code>' + html.escape(recipe['meta']['name']) + '</code>. '
             'Description and data type come from the child object itself and will '
             'appear once its FHX is imported.</div>')
    h.append('<input class="alias-filter" id="rpFilter" placeholder="Filter parameters\u2026" oninput="rpFilterGrid(this)">')
    h.append('<div class="rp-gridwrap"><table class="fbd-table rp-grid"><thead><tr>'
             '<th>Name</th><th>Type</th><th>Parameter Value</th><th>Origin</th><th>Group</th>'
             '</tr></thead><tbody>')
    for p in s.get('params', []):
        if p.get('origin') == 'DEFERRED':
            pv = ('<code class="rp-upref">' + html.escape(p.get('deferred_to') or p['name'])
                  + '</code> <span class="rp-uparrow">\u2191 ' + html.escape(recipe['meta']['name']) + '</span>')
        else:
            pv = '<span class="rp-none">(constant \u2014 value set on the child object)</span>'
        h.append('<tr class="rp-row">'
                 '<td><code>' + html.escape(p['name']) + '</code></td>'
                 '<td>Recipe Parameter</td>'
                 '<td>' + pv + '</td>'
                 '<td>' + html.escape(p.get('origin', '')) + '</td>'
                 '<td>' + html.escape(p.get('group', '')) + '</td>'
                 '</tr>')
    h.append('</tbody></table></div></div>')
    return '\n'.join(h)


def build_step_views(recipes):
    """For every recipe, build the derived parameter view for each of its step
    instances (loaded or not). Keyed 'PARENT||STEP' for the client map."""
    out = {}
    for r in recipes:
        proc = r.get('procedure') or {}
        for sn, s in (proc.get('steps') or {}).items():
            if sn in ('START', 'END') or not s.get('params'):
                continue
            v = build_step_param_html(r, sn)
            if v:
                out[r['meta']['name'] + '||' + sn] = v
    return out


def build_deferrals_rows(recipe):
    """Flatten every step's parameters into audit rows matching the deferral-check
    format: one row per (step, parameter), with the deferred-to name when the
    parameter is deferred (blank otherwise). Order follows the FHX step order, and
    within a step the parameter declaration order — this is the same shape as a
    hand audit of 'Contents of <object>' in DeltaV Recipe Studio, just automatic
    and covering every step (not just the ones manually reviewed).
    Returns [{'step': step_name, 'definition': ..., 'param': ..., 'deferred_to': ...}]."""
    rows = []
    proc = recipe.get('procedure') or {}
    for sn, s in (proc.get('steps') or {}).items():
        if sn in ('START', 'END') or not s.get('params'):
            continue
        for p in s['params']:
            deferred_to = p.get('deferred_to', '') if p.get('origin') == 'DEFERRED' else ''
            rows.append({'step': sn, 'definition': s.get('definition', ''),
                        'param': p['name'], 'deferred_to': deferred_to})
    return rows


def build_deferrals_html(recipe, rows=None):
    """Render the deferrals as a scannable, grouped view: one collapsible subcard per
    step showing which object it instantiates and its parameters, with deferred ones
    called out visually. This is the in-app working view — an Excel export of the
    same data is one click away for sharing, but the grid isn't the primary artifact."""
    rows = rows if rows is not None else build_deferrals_rows(recipe)
    if not rows:
        return ''
    # group rows by step, preserving file order
    steps = []
    seen = {}
    for r in rows:
        if r['step'] not in seen:
            seen[r['step']] = {'step': r['step'], 'definition': r['definition'], 'params': []}
            steps.append(seen[r['step']])
        seen[r['step']]['params'].append(r)

    n_params = len(rows)
    n_deferred = sum(1 for r in rows if r['deferred_to'])
    pct = round(100 * n_deferred / n_params) if n_params else 0

    h = ['<div class="card" style="max-width:none"><h3>Deferrals '
         '<span style="font-weight:400;color:var(--ink-3);font-size:12px">'
         '\u2014 included in the PFC report (.xlsx)</span></h3>']
    h.append('<div class="def-summary">'
             '<span class="def-chip">' + str(len(steps)) + ' steps</span>'
             '<span class="def-chip">' + str(n_params) + ' parameters</span>'
             '<span class="def-chip def-chip-hl">' + str(n_deferred) + ' deferred (' + str(pct) + '%)</span>'
             '</div>')
    h.append('<div class="def-toolbar">'
             '<input class="alias-filter" id="defFilter" placeholder="Filter by step or parameter\u2026" oninput="defFilterGrid(this)" style="flex:1">'
             '<label class="def-toggle"><input type="checkbox" id="defOnlyDeferred" onchange="defFilterGrid(this.closest(\'.card\').querySelector(\'#defFilter\'))"> deferred only</label>'
             '<span class="link" onclick="defSetAll(true)">expand all</span>'
             '<span class="link" onclick="defSetAll(false)">collapse all</span>'
             '</div>')
    h.append('<div id="defList">')
    for sgrp in steps:
        ndef = sum(1 for p in sgrp['params'] if p['deferred_to'])
        h.append('<div class="subcard def-step-card collapsed" data-defcount="' + str(ndef) + '">')
        h.append('<h4><code>' + html.escape(sgrp['step']) + '</code>'
                 + (' <span class="def-def">\u2192 ' + html.escape(sgrp['definition']) + '</span>' if sgrp['definition'] else '')
                 + ' <span class="def-count">' + str(len(sgrp['params'])) + ' params'
                 + (', ' + str(ndef) + ' deferred' if ndef else '') + '</span></h4>')
        h.append('<div class="def-plist">')
        for p in sgrp['params']:
            if p['deferred_to']:
                h.append('<div class="def-prow def-prow-deferred"><code>' + html.escape(p['param']) + '</code>'
                         ' <span class="def-arrow">\u2192</span> <code class="rp-upref">' + html.escape(p['deferred_to'])
                         + '</code></div>')
            else:
                h.append('<div class="def-prow"><code>' + html.escape(p['param']) + '</code>'
                         ' <span class="def-const">constant</span></div>')
        h.append('</div></div>')
    h.append('</div></div>')
    return '\n'.join(h)


def build_deferrals_xlsx(recipe_name, rows):
    """Build an .xlsx matching the audit format: Unit Procedure | Parameter Name |
    Deferred Parameter, with quoted values and the step name shown only on the
    first row of its group (forward-fill visual style) — the same layout as a
    manual DeltaV Recipe Studio deferral audit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deferrals'
    headers = ['Unit Procedure', 'Parameter Name', 'Deferred Parameter']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='EEF2F7')

    last_step = None
    for r in rows:
        step_val = ('STEP NAME="' + r['step'] + '"') if r['step'] != last_step else None
        last_step = r['step']
        ws.append([step_val, '"' + r['param'] + '"',
                  ('"' + r['deferred_to'] + '"') if r['deferred_to'] else None])
    for i, w in enumerate((32, 30, 30), start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pfc_report_xlsx(recipe, formulas=None):
    """Converter-style structured report for one recipe object. Four sheets:
      Overview   — identity, batch sizing, and the formulas present in the export
      Parameters — the full grid, with ONE COLUMN PER FORMULA so value sets read
                   side-by-side (e.g. MEDIA AND 20 vs MEDIA AND 25) instead of
                   flipping a selector
      Procedure  — the PFC as an ordered walk: each step (definition, layer, unit
                   alias, param/deferral counts) followed by its outgoing
                   transitions with expressions
      Deferrals  — the per-step deferral audit (same layout as the single export)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    formulas = formulas or []
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill('solid', start_color='EEF2F7')
    step_fill = PatternFill('solid', start_color='F1F5F9')
    defer_font = Font(color='B45309', bold=True)

    def _head(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
        ws.freeze_panes = 'A2'

    wb = Workbook()

    # ── Overview ──
    ws = wb.active
    ws.title = 'Overview'
    meta = recipe['meta']
    ws.append(['Recipe', meta.get('name', '')])
    ws['A1'].font = hdr_font
    for k, label in [('type', 'Type'), ('description', 'Description'),
                     ('author', 'Author'), ('version', 'Version'),
                     ('product_code', 'Product code'), ('product_name', 'Product name'),
                     ('batch_units', 'Batch units'), ('default_batch_size', 'Default batch size'),
                     ('min_batch_size', 'Minimum batch size'), ('max_batch_size', 'Maximum batch size')]:
        v = meta.get(k, '')
        if v:
            ws.append([label, v])
    proc = recipe.get('procedure') or {}
    ws.append([])
    ws.append(['Steps', len([s for s in (proc.get('steps') or {}) if s not in ('START', 'END')])])
    ws.append(['Transitions', len(proc.get('transitions') or {})])
    ws.append(['Formula parameters', len(recipe.get('params', []))])
    if formulas:
        ws.append([])
        ws.append(['Formulas in export'])
        ws['A' + str(ws.max_row)].font = hdr_font
        ws.append(['Name', 'Released', 'Values', 'Description'])
        for c in ws[ws.max_row]:
            c.font = hdr_font
            c.fill = hdr_fill
        for f in formulas:
            ws.append([f['name'], 'Yes' if f.get('released') else '',
                      len(f.get('values', {})), f.get('description', '')])
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 46

    # ── Parameters (formulas side-by-side) ──
    ws = wb.create_sheet('Parameters')
    base_cols = ['Name', 'Type', 'Description', 'Parameter Type', 'Filtering', 'Group', 'Locked']
    fcols = [f['name'] for f in formulas]
    _head(ws, base_cols + fcols)
    fvals = {f['name']: f.get('values', {}) for f in formulas}
    for p in recipe.get('params', []):
        row = [p['name'], _fhx_type_label(p.get('connection', '')), p.get('description', ''),
               p.get('param_type', ''), p.get('filtering', ''), p.get('group', ''),
               'Yes' if p.get('locked') else '']
        for fn in fcols:
            row.append(fvals.get(fn, {}).get(p['name'], ''))
        ws.append(row)
    widths = [30, 16, 44, 14, 10, 12, 8] + [16] * len(fcols)
    for i, w in enumerate(widths, start=1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = w

    # ── Procedure (PFC walk) ──
    ws = wb.create_sheet('Procedure')
    _head(ws, ['Kind', 'Name', 'Definition / To-step', 'Layer', 'Unit alias',
               'Params', 'Deferred', 'Expression'])
    steps = proc.get('steps') or {}
    trans = proc.get('transitions') or {}
    step_outs = {}
    for s, t in (proc.get('s2t') or []):
        step_outs.setdefault(s, []).append(t)
    t2s = {}
    for t, s in (proc.get('t2s') or []):
        t2s.setdefault(t, []).append(s)
    for sn, s in steps.items():
        if sn in ('START', 'END'):
            r = ws.max_row + 1
            ws.append(['STEP', sn, '', '', '', '', '', ''])
            for c in ws[r]:
                c.fill = step_fill
                c.font = hdr_font
        else:
            r = ws.max_row + 1
            ws.append(['STEP', sn, s.get('definition', ''), _layer_label(s.get('definition', '')),
                      s.get('unit_alias', ''), len(s.get('params', [])),
                      len(s.get('deferred', [])), ''])
            for c in ws[r]:
                c.fill = step_fill
            ws.cell(row=r, column=2).font = hdr_font
            if s.get('deferred'):
                ws.cell(row=r, column=7).font = defer_font
        for tn in step_outs.get(sn, []):
            td = trans.get(tn, {})
            to = ', '.join(t2s.get(tn, [])) or '?'
            rr = ws.max_row + 1
            ws.append(['  transition', tn, '\u2192 ' + to, '', '', '', '',
                      td.get('expr', '') or '(state transition)'])
            ws.cell(row=rr, column=8).alignment = Alignment(wrap_text=True)
    for i, w in enumerate((12, 30, 30, 14, 16, 8, 9, 70), start=1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = w

    # ── Deferrals ──
    # A proper deferral audit, not a raw dump: a summary line, then one banded block
    # per step (the object it instantiates), each parameter on its own row with its
    # group, and DEFERRED parameters called out with their target. Only deferred rows
    # carry the highlight so a reviewer's eye lands on what actually gets bound later.
    ws = wb.create_sheet('Deferrals')
    defer_rows = build_deferrals_rows(recipe)
    n_all = len(defer_rows)
    n_def = sum(1 for r in defer_rows if r['deferred_to'])
    # group params by step in file order + collect each param's group/type from the step
    proc = recipe.get('procedure') or {}
    step_params = {sn: {p['name']: p for p in (s.get('params') or [])}
                   for sn, s in (proc.get('steps') or {}).items()}

    ws.append(['Deferral audit \u2014 ' + meta.get('name', '')])
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws.append([f'{n_def} of {n_all} step parameter(s) are deferred '
               f'({round(100 * n_def / n_all) if n_all else 0}%). '
               f'Deferred rows are highlighted and show the parameter they bind to.'])
    ws['A2'].font = Font(italic=True, size=9, color='666666')
    ws.append([])
    hdr_row = ws.max_row + 1
    ws.append(['Step', 'Instantiates', 'Parameter', 'Group', 'Deferred?', 'Binds To'])
    for c in ws[hdr_row]:
        c.font = hdr_font
        c.fill = step_fill
    ws.freeze_panes = f'A{hdr_row + 1}'

    band = PatternFill('solid', fgColor='F4F7FB')
    defer_fill = PatternFill('solid', fgColor='FFF3CD')
    last_step = None
    band_on = False
    for r in defer_rows:
        new_step = r['step'] != last_step
        if new_step:
            band_on = not band_on
        pinfo = step_params.get(r['step'], {}).get(r['param'], {})
        grp = pinfo.get('group', '')
        is_def = bool(r['deferred_to'])
        rr = ws.max_row + 1
        ws.append([r['step'] if new_step else None,
                   r['definition'] if new_step else None,
                   r['param'], grp,
                   'DEFERRED' if is_def else '', r['deferred_to'] or None])
        for c in ws[rr]:
            if is_def:
                c.fill = defer_fill
            elif band_on:
                c.fill = band
        if new_step:
            ws.cell(row=rr, column=1).font = hdr_font
        if is_def:
            ws.cell(row=rr, column=5).font = defer_font
        last_step = r['step']
    for i, w in enumerate((30, 30, 30, 16, 12, 30), start=1):
        col = ws.cell(row=hdr_row, column=i).column_letter
        ws.column_dimensions[col].width = w

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_deferrals_all_xlsx(recipes):
    """One workbook, one sheet per recipe object (like the Converter's document
    outputs) — each sheet in the same audit layout as the single-recipe export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    for rec in recipes:
        rows = build_deferrals_rows(rec)
        if not rows:
            continue
        title = re.sub(r'[\\/*?:\[\]]', '_', rec['meta']['name'])[:31] or 'Recipe'
        ws = wb.create_sheet(title=title)
        ws.append(['Unit Procedure', 'Parameter Name', 'Deferred Parameter'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='EEF2F7')
        last_step = None
        for r in rows:
            step_val = ('STEP NAME="' + r['step'] + '"') if r['step'] != last_step else None
            last_step = r['step']
            ws.append([step_val, '"' + r['param'] + '"',
                      ('"' + r['deferred_to'] + '"') if r['deferred_to'] else None])
        for i, w in enumerate((32, 30, 30), start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'
    if not wb.sheetnames:
        ws = wb.create_sheet(title='Deferrals')
        ws.append(['No recipe step parameters found in this import.'])

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fhx_param_type(t):
    """Map a FORMULA_PARAMETER TYPE code to the 'Parameter Type' label DeltaV shows."""
    t = (t or '').upper()
    m = {
        'BATCH_PARAMETER_REAL': 'Real',
        'BATCH_PARAMETER_INTEGER': 'Integer',
        'BATCH_PARAMETER_BOOLEAN': 'Boolean',
        'BATCH_PARAMETER_STRING': 'String',
        'BATCH_PARAMETER_ENUMERATION': 'Named Set',
        'ENUMERATION_VALUE': 'Named Set',
        'NAMED_SET': 'Named Set',
    }
    if t in m:
        return m[t]
    if 'REAL' in t:
        return 'Real'
    if 'INT' in t:
        return 'Integer'
    if 'BOOL' in t:
        return 'Boolean'
    if 'ENUM' in t or 'NAMED' in t:
        return 'Named Set'
    if 'STRING' in t:
        return 'String'
    return t.title() if t else ''


def _fhx_filtering(connection):
    """Map CONNECTION to the 'Filtering' column DeltaV shows (Process Input / Input)."""
    c = (connection or '').upper()
    if c == 'INPUT':
        return 'Input'
    if c == 'OUTPUT':
        return 'Output'
    if c:
        return c.title()
    return ''


def _fhx_type_label(connection, is_step_param=False):
    """The 'Type' column: 'Recipe Parameter' / 'Batch Parameter' / 'Parameter'."""
    c = (connection or '').upper()
    if is_step_param:
        return 'Recipe Parameter'
    if c == 'INPUT':
        return 'Batch Parameter'
    return 'Parameter'


def _layer_label(definition):
    """Infer the recipe layer of a step's definition from DeltaV naming conventions:
    _UP = unit procedure, _OP = operation, _PH/PHASE = phase. Falls back to a generic
    label so the hierarchy Procedure > Unit Procedure > Operation > Phase reads clearly."""
    d = (definition or '').upper()
    if d.endswith('_UP') or '_UP' in d:
        return 'unit procedure'
    if d.endswith('_OP') or '_OP' in d:
        return 'operation'
    if d.endswith('_PH') or 'PHASE' in d:
        return 'phase'
    return 'step definition'


def _safe_id(s):
    return re.sub(r'[^A-Za-z0-9]', '_', s or 'x')


RECIPE_CSS = """
.rflow{display:flex;flex-direction:column;gap:0;font-size:13px}
.rf-step{border:1px solid var(--border,#e2e8f0);border-radius:9px;padding:9px 12px;background:var(--surface,#fff);margin:3px 0}
.rf-start{background:#0f172a;color:#fff;font-weight:700;text-align:center;max-width:120px}
.rf-name{font-weight:650;color:var(--ink,#16202c)}
.rf-init{font-size:10px;font-weight:700;background:#dcfce7;color:#166534;padding:1px 6px;border-radius:5px;margin-left:6px}
.rf-def{color:var(--ink-3,#64748b);font-size:12px;margin-top:2px}
.rf-trans{padding:3px 0 3px 22px;position:relative;color:#6d28d9;font-size:12px}
.rf-tname{font-weight:600}
.rf-tgt{color:var(--ink-3,#64748b)}
.rf-expr{font-family:ui-monospace,Menlo,monospace;font-size:11px;color:#475569;background:#f8fafc;border:1px solid var(--border,#e2e8f0);border-radius:6px;padding:4px 8px;margin:3px 0 3px 0;white-space:pre-wrap;word-break:break-word}
.rp-group{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:#64748b;margin:12px 0 6px}
.rp-gridwrap{overflow-x:auto;max-height:560px;overflow-y:auto;border:1px solid var(--border,#e2e8f0);border-radius:8px}
.rp-grid{font-size:12px;min-width:900px}
.rp-grid th{position:sticky;top:0;background:#f1f5f9;z-index:1}
.rp-grid .rp-desc{color:#475569;max-width:280px}
.rp-none{color:#cbd5e1}
.rp-upref{color:#2563eb}
.rp-uparrow{font-size:10px;color:#94a3b8;font-weight:600}
.rp-val{color:#0f766e;font-weight:600}
.rp-formula-bar{display:flex;align-items:center;gap:8px;margin-bottom:9px;font-size:12.5px;color:#334155;flex-wrap:wrap}
.rp-editbtn{background:var(--surface-2,#f1f5f9);border:1px solid var(--border,#e2e8f0);border-radius:7px;padding:5px 11px;font-size:12.5px;font-weight:600;cursor:pointer;color:var(--ink,#16202c)}
.rp-editbtn:hover{border-color:var(--accent,#2563eb);color:var(--accent,#2563eb)}
.rp-editbtn-on{background:#fef2f2;border-color:#fca5a5;color:#b91c1c}
.rp-editin{width:100%;box-sizing:border-box;padding:3px 6px;border:1px solid var(--accent,#2563eb);border-radius:5px;font-family:'IBM Plex Mono',ui-monospace,monospace;font-size:12px;background:#fffbeb}
.rp-grid.rp-editing .rp-valcell.rp-noedit{opacity:.5}
.rp-editstatus{flex-basis:100%;margin-top:6px;font-size:12px;color:var(--ink-2,#475569)}
.rp-editstatus .link{margin-left:8px;font-weight:600}
.rp-formula-bar select{background:var(--surface-2,#f1f5f9);color:var(--ink,#16202c);border:1px solid var(--border,#e2e8f0);border-radius:7px;padding:5px 10px;font-size:13px;font-weight:600}
.def-summary{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap}
.rp-formula-hint{font-size:11px;color:#94a3b8}
.rstep-note{font-size:12px;color:#64748b;background:#f8fafc;border:1px solid #e2e8f0;border-radius:7px;padding:7px 11px;margin-bottom:10px}
.def-chip{font-size:11px;font-weight:600;background:#f1f5f9;color:#475569;padding:3px 10px;border-radius:20px}
.def-chip-hl{background:#fef3c7;color:#92400e}
.def-toolbar{display:flex;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap}
.def-toolbar .alias-filter{max-width:320px}
.def-toggle{display:flex;align-items:center;gap:5px;font-size:12px;color:#475569;white-space:nowrap}
.def-toolbar .link{font-size:12px;white-space:nowrap}
.def-step-card{padding:9px 12px}
.def-step-card h4{font-size:13px}
.def-def{font-size:11px;color:#7c3aed;font-weight:600}
.def-count{font-size:10.5px;color:#94a3b8;font-weight:500;margin-left:auto}
.def-plist{display:flex;flex-direction:column;gap:3px;padding-left:4px}
.def-prow{font-size:12px;color:#334155;padding:3px 0}
.def-prow-deferred{color:#1e293b}
.def-arrow{color:#94a3b8;margin:0 3px}
.def-const{font-size:10px;color:#94a3b8;font-style:italic;margin-left:6px}
.def-step-card.def-hide{display:none}
.def-prow.def-hide{display:none}
.rf-alias{font-size:10px;font-weight:600;background:#e0f2fe;color:#075985;padding:1px 7px;border-radius:5px;margin-left:6px}
.rf-defcount{color:#b45309;font-weight:600}
.rf-drill{cursor:pointer;color:#7c3aed}
.rf-drill:hover{text-decoration:underline}
.rf-defers{margin:5px 0 2px 14px;padding:7px 10px;background:#fffbeb;border:1px solid #fde68a;border-radius:7px}
.rf-defhdr{font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.03em;color:#b45309;margin-bottom:5px}
.rf-defrow{font-size:12px;color:#475569;padding:1px 0}
.rf-defgrp{font-size:10px;color:#94a3b8;margin-left:5px}
.rf-defval{color:#16a34a;font-weight:600;font-size:11px}
.rf-defup{color:#2563eb;font-weight:600;font-size:11px}
.rf-defun{color:#dc2626;font-size:10.5px;font-style:italic}
.pfc-wrap{overflow:auto;border:1px solid var(--border,#e2e8f0);border-radius:9px;background:#fff;padding:10px;max-height:520px;cursor:grab;position:relative}
.pfc-overlay{position:fixed;inset:0;background:rgba(15,23,42,.55);z-index:60;display:grid;place-items:center;padding:26px}
.pfc-ov-card{background:var(--surface,#fff);border-radius:14px;width:min(1200px,94vw);height:min(88vh,900px);display:flex;flex-direction:column;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.35)}
.pfc-ov-h{display:flex;align-items:center;gap:12px;padding:12px 18px;border-bottom:1px solid var(--border);flex-wrap:wrap}
.pfc-ov-hint{font-size:11px;color:var(--ink-3)}
.pfc-ov-x{background:none;border:none;font-size:22px;line-height:1;cursor:pointer;color:var(--ink-3);padding:0 4px}
.pfc-ov-x:hover{color:var(--ink)}
.pfc-ov-body{flex:1;overflow:hidden;padding:14px;display:flex;min-height:0}
/* #3: diagram on the left, a drag bar, then the detail panel pinned on the RIGHT */
.pfc-ov-split{flex:1;display:flex;min-width:0;min-height:0;gap:0}
.pfc-ov-diagram{flex:1 1 auto;min-width:0;overflow:hidden;display:flex}
.pfc-ov-diagram .pfc-wrap{flex:1;max-height:none;height:100%}
.pfc-ov-divider{flex:0 0 6px;cursor:col-resize;background:var(--border,#e2e8f0);border-radius:3px;margin:0 4px}
.pfc-ov-divider:hover{background:var(--accent,#7c3aed)}
.pfc-ov-panel{flex:0 0 320px;min-width:220px;overflow:auto;align-self:stretch;margin-top:0}
.pfc-ov-empty{padding:40px;text-align:center;color:var(--ink-3)}
.pfc-wrap.panning{cursor:grabbing}
.pfc-zoomlayer{transform-origin:0 0;transition:transform .05s linear;display:inline-block}
.pfc-toolbar{display:flex;align-items:center;gap:5px;margin-bottom:7px}
.pfc-zbtn{width:26px;height:26px;border:1px solid var(--border,#e2e8f0);background:var(--surface,#fff);border-radius:6px;cursor:pointer;font-size:14px;line-height:1;color:var(--ink,#16202c);display:inline-flex;align-items:center;justify-content:center}
.pfc-zbtn:hover{border-color:var(--accent,#7c3aed);color:var(--accent,#7c3aed)}
.pfc-hint2{font-size:11px;color:var(--ink-3,#94a3b8);margin-left:6px}
.pfc-svg{display:block}
.pfc-step rect{transition:stroke .12s}
.pfc-drillable{cursor:pointer}
.pfc-drillable:hover rect{stroke:#7c3aed;stroke-width:2.5}
.pfc-drillable:hover text{fill:#7c3aed}
.pfc-trans:hover rect{fill:#f5f3ff}
.pfc-panel{margin-top:10px;padding:10px 12px;border:1px solid var(--border,#e2e8f0);border-radius:8px;background:#f8fafc;min-height:24px}
.pfc-layout{display:flex;gap:0;align-items:stretch}
.pfc-layout .pfc-wrap{flex:1 1 auto;min-width:0;order:1}
.pfc-divider{flex:0 0 6px;cursor:col-resize;background:var(--border,#e2e8f0);border-radius:3px;margin:0 6px;order:2}
.pfc-divider:hover{background:var(--accent,#7c3aed)}
.pfc-layout .pfc-panel{margin-top:0;flex:0 0 320px;min-width:200px;max-height:520px;overflow:auto;align-self:stretch;order:3}
@media(max-width:900px){.pfc-layout{flex-direction:column}.pfc-divider{display:none}.pfc-layout .pfc-panel{flex:none;width:100%;max-height:none;order:3}}
.pfc-hint{color:#94a3b8;font-size:12px}
.pfc-tname{font-weight:700;color:#7c3aed;font-size:12px;margin-bottom:4px}
.pfc-texpr{font-family:ui-monospace,Menlo,monospace;font-size:11.5px;color:#334155;white-space:pre-wrap;word-break:break-word}
.pfc-step{cursor:pointer}
.pfc-step:hover rect{stroke:#7c3aed}
.pfc-step.sel rect{stroke:#7c3aed;stroke-width:2.5}
.pfc-trans.sel rect{stroke-width:2.5}
.pfc-sdef{font-size:11.5px;color:#475569;margin-bottom:6px}
.pfc-sdrill{margin:4px 0 8px}
.pfc-sdrill .link{color:#7c3aed;font-weight:600;font-size:12px;cursor:pointer}
.pfc-sdrill .link:hover{text-decoration:underline}
.pfc-sdrill-ghost{font-size:11.5px;color:#94a3b8}
.pfc-sparams{width:100%;border-collapse:collapse;font-size:11.5px;margin-top:4px}
.pfc-sparams th{text-align:left;padding:3px 8px;color:#64748b;font-size:10px;text-transform:uppercase;letter-spacing:.03em;border-bottom:1px solid var(--border,#e2e8f0)}
.pfc-sparams td{padding:3px 8px;border-bottom:1px solid var(--border,#eef2f6)}
.pfc-sparams tr:last-child td{border-bottom:0}
.pfc-sparams .pfc-pdef td{background:#fffbeb}
.pfc-sparams code{font-size:11px;color:#b45309}
.pfc-shead{margin-bottom:6px}
.pfc-tags{display:flex;flex-wrap:wrap;gap:4px;margin-top:4px}
.pfc-tag{font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;padding:1px 7px;border-radius:10px;background:#eef2f7;color:#475569}
.pfc-tag-init{background:#dcfce7;color:#166534}
.pfc-tag-def{background:#fef3c7;color:#b45309}
.pfc-seq{margin:6px 0 8px;padding:6px 9px;background:#f1f5f9;border-radius:6px;font-size:11px}
.pfc-seqrow{padding:1px 0;color:#475569}
.pfc-seqlbl{display:inline-block;width:32px;color:#94a3b8;font-size:10px;text-transform:uppercase}
.pfc-seqtrans{cursor:pointer;color:#7c3aed}
.pfc-seqtrans:hover{text-decoration:underline}
.pfc-plabel{font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.03em;color:#64748b;margin:6px 0 3px}
"""

